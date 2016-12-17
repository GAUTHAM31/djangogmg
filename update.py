import os,sys,django
sys.path.append(os.path.join(os.path.dirname(__file__), '/root/gmg_1'))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "gmg_1.settings")
from django.conf import settings
from django.db import models
django.setup()
from django.contrib.auth.models import User
from attendance.models import employee
from openpyxl import load_workbook
#------------------Code----------------------------------------------------


wb = load_workbook(filename='/root/Book1.xlsx')
sheet  = wb.active
employee.objects.all().update(present=0)#sets all absent at first
#loop to parse excel file
for row in range(2, sheet.max_row + 1):
	employee.objects.all().filter(eid=sheet['B'+str(row)].value).update(present=1)


	

 
